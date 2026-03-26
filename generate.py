#!/usr/bin/env python3
"""
generate.py — Generate a self-contained DataTables HTML catalog from an Excel file.

Usage:
    python3 generate.py data/oer-catalog-2025-2026.xlsx

The output HTML file is written to the current directory with the same base name
as the input file (e.g., oer-catalog-2025-2026.html).
"""

import sys
import os
import re
import html
import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    sys.exit("Error: openpyxl is required. Install it with: pip install openpyxl")


FILENAME_PATTERN = re.compile(r'^oer-catalog-\d{4}-\d{4}\.xlsx$')
REQUIRED_COLUMNS = ["OER Title", "Link"]
FILTER_COLUMNS = ["Campus", "Type", "Discipline", "Platform"]

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title}</title>
  <link rel="stylesheet" href="https://cdn.datatables.net/1.13.8/css/jquery.dataTables.min.css">
  <style>
    *, *::before, *::after {{ box-sizing: border-box; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      font-size: 14px;
      color: #222;
      margin: 0;
      padding: 16px;
      background: #fff;
    }}
    h1 {{
      font-size: 1.25rem;
      font-weight: 600;
      margin: 0 0 16px;
    }}
    .filters {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-bottom: 14px;
      align-items: center;
    }}
    .filters label {{
      display: flex;
      flex-direction: column;
      gap: 3px;
      font-size: 12px;
      font-weight: 600;
      color: #555;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .filters select {{
      font-size: 13px;
      padding: 5px 8px;
      border: 1px solid #767676;
      border-radius: 4px;
      background: #fff;
      cursor: pointer;
      min-width: 160px;
    }}
    .filters select:focus {{
      outline: 2px solid #4a90d9;
      outline-offset: 1px;
    }}
    .table-wrapper {{
      width: 100%;
      overflow-x: auto;
    }}
    #oer-catalog {{
      width: 100% !important;
    }}
    #oer-catalog thead th {{
      white-space: nowrap;
    }}
    #oer-catalog tbody td {{
      vertical-align: top;
      line-height: 1.4;
    }}
    #oer-catalog a {{
      color: #1a5494;
      text-decoration: none;
    }}
    #oer-catalog a:hover {{
      text-decoration: underline;
    }}
    #oer-catalog a:focus {{
      outline: 2px solid #4a90d9;
      outline-offset: 2px;
    }}
    .dataTables_wrapper .dataTables_filter input {{
      border: 1px solid #767676;
      border-radius: 4px;
      padding: 4px 8px;
    }}
    .sr-only {{
      position: absolute;
      width: 1px;
      height: 1px;
      padding: 0;
      margin: -1px;
      overflow: hidden;
      clip: rect(0, 0, 0, 0);
      white-space: nowrap;
      border: 0;
    }}
  </style>
</head>
<body>
  <h1 id="catalog-title">{title}</h1>
  <div class="filters">
{filter_selects}
  </div>
  <div class="table-wrapper">
    <table id="oer-catalog" class="display" style="width:100%" aria-labelledby="catalog-title">
      <thead>
        <tr>
{header_cells}
        </tr>
      </thead>
      <tbody>
{body_rows}
      </tbody>
    </table>
  </div>

  <script src="https://cdn.jsdelivr.net/npm/iframe-resizer@4/js/iframeResizer.contentWindow.min.js"></script>
  <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
  <script src="https://cdn.datatables.net/1.13.8/js/jquery.dataTables.min.js"></script>
  <script>
  $(function () {{
    var table = $('#oer-catalog').DataTable({{
      pageLength: 25,
      lengthMenu: [10, 25, 50, 100],
      order: [[1, 'asc']],
      columnDefs: [{{ targets: '_all', defaultContent: '' }}]
    }});

    // Column index map for dropdown filters
    var filterCols = {filter_col_map};

    // Custom search function for dropdowns
    $.fn.dataTable.ext.search.push(function (settings, data) {{
      for (var col in filterCols) {{
        var val = $('#filter-' + col.toLowerCase()).val();
        if (val && data[filterCols[col]] !== val) {{
          return false;
        }}
      }}
      return true;
    }});

    // Trigger redraw on dropdown change
    $('.filters select').on('change', function () {{
      table.draw();
    }});
  }});
  </script>
</body>
</html>
"""


def make_select(col_name, options):
    """Build a <label>/<select> block for a filter dropdown."""
    col_id = col_name.lower()
    opt_tags = ['      <option value="">All {}</option>'.format(col_name)]
    for opt in sorted(options):
        esc = html.escape(str(opt), quote=True)
        opt_tags.append('      <option value="{0}">{0}</option>'.format(esc))
    return (
        '    <label>{col}\n'
        '      <select id="filter-{col_id}">\n'
        '{opts}\n'
        '      </select>\n'
        '    </label>'
    ).format(col=col_name, col_id=col_id, opts='\n'.join(opt_tags))


def cell_value(row, index):
    """Safely retrieve a cell value from a row, returning None if out of bounds."""
    if index < len(row):
        return row[index]
    return None


def generate(excel_path: str) -> str:
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except InvalidFileException:
        sys.exit("Error: '{}' is not a valid Excel file.".format(excel_path))
    except Exception as e:
        sys.exit("Error: could not open '{}': {}".format(excel_path, e))

    try:
        if "Catalog" not in wb.sheetnames:
            sys.exit("Error: sheet 'Catalog' not found. Available sheets: {}".format(
                ", ".join(wb.sheetnames)))
        ws = wb["Catalog"]

        rows = [row for row in ws.iter_rows(values_only=True)]
        if not rows:
            sys.exit("Error: Catalog sheet is empty.")

        # Normalise headers: strip regular and non-breaking whitespace
        raw_headers = rows[0]
        if all(h is None for h in raw_headers):
            sys.exit("Error: header row is empty — check that row 1 contains column names.")
        headers = [str(h).replace('\xa0', '').strip() if h is not None else "" for h in raw_headers]

        # Detect duplicate column names
        seen = {}
        for i, h in enumerate(headers):
            if h in seen:
                sys.exit("Error: duplicate column name '{}' at positions {} and {}.".format(
                    h, seen[h] + 1, i + 1))
            seen[h] = i

        # Verify required columns are present
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            sys.exit("Error: required column(s) not found: {}.".format(", ".join(missing)))

        link_idx = headers.index("Link")

        data_rows = rows[1:]
        # Filter out fully empty rows
        data_rows = [r for r in data_rows if not all(v is None or str(v).strip() == "" for v in r)]

        if not data_rows:
            print("Warning: no data rows found — the generated table will be empty.")

        # Visible columns: all except Link
        visible_headers = [h for h in headers if h != "Link"]

        # Build filter column index map (using visible column indices)
        filter_col_map = {}
        for col in FILTER_COLUMNS:
            if col in visible_headers:
                filter_col_map[col] = visible_headers.index(col)

        # Collect unique values for each filter column
        filter_options = {col: set() for col in FILTER_COLUMNS if col in headers}
        for row in data_rows:
            for col in filter_options:
                idx = headers.index(col)
                val = cell_value(row, idx)
                if val is not None and str(val).strip():
                    filter_options[col].add(str(val).strip())

        # Build filter selects HTML
        filter_selects = "\n".join(
            make_select(col, filter_options[col])
            for col in FILTER_COLUMNS
            if col in filter_options
        )

        # Build header cells
        header_cells = "\n".join(
            '          <th scope="col">{}</th>'.format(html.escape(h))
            for h in visible_headers
        )

        # Build body rows
        body_rows_list = []
        for row in data_rows:
            raw_link = cell_value(row, link_idx)
            link = str(raw_link).strip() if raw_link is not None else ""
            cells = []
            for i, h in enumerate(headers):
                if h == "Link":
                    continue
                val = cell_value(row, i)
                cell_text = str(val).strip() if val is not None else ""
                if h == "OER Title" and link:
                    cell_html = (
                        '<a href="{href}" target="_blank" rel="noopener">'
                        '{text}<span class="sr-only"> (opens in new tab)</span>'
                        '</a>'
                    ).format(
                        href=html.escape(link, quote=True),
                        text=html.escape(cell_text or "(Untitled)"),
                    )
                else:
                    cell_html = html.escape(cell_text)
                cells.append("          <td>{}</td>".format(cell_html))
            body_rows_list.append("        <tr>\n{}\n        </tr>".format("\n".join(cells)))

        body_rows = "\n".join(body_rows_list)

        stem = Path(excel_path).stem
        title = stem.replace("-", " ").title().replace("Oer", "OER")

        return HTML_TEMPLATE.format(
            title=title,
            filter_selects=filter_selects,
            header_cells=header_cells,
            body_rows=body_rows,
            filter_col_map=json.dumps(filter_col_map),
        )
    finally:
        wb.close()


def main():
    if len(sys.argv) != 2:
        sys.exit("Usage: python3 generate.py <path/to/excel-file.xlsx>")

    excel_path = sys.argv[1]

    if not os.path.isfile(excel_path):
        sys.exit("Error: file not found: {}".format(excel_path))

    filename = Path(excel_path).name
    if not FILENAME_PATTERN.match(filename):
        sys.exit(
            "Error: filename '{}' does not match the required format.\n"
            "Expected: oer-catalog-YYYY-YYYY.xlsx (e.g. oer-catalog-2025-2026.xlsx)".format(filename)
        )

    output_path = Path(excel_path).stem + ".html"

    html_content = generate(excel_path)

    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)
    except OSError as e:
        sys.exit("Error: could not write '{}': {}".format(output_path, e))

    print("Generated: {}".format(output_path))


if __name__ == "__main__":
    main()
